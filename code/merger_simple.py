import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

SIMILARITY_THRESHOLD = 0.20
MIN_MATCHES = 5
MAX_MATCHES = 15
FILTER_BY_THEME = True
RANDOM_SEED = None
CURRENT_FILE = "../data/artemis/artemis_data_numeric.xlsx"
PAST_FILE = "../data/comparison_data/previous_projects_data_cleaned.xlsx"
OUTPUT_FILE = "../data/interval_data/interval_analysis_simple.xlsx"


def load_data():
    """Load current and past project data."""
    current = pd.read_excel(CURRENT_FILE)
    past = pd.read_excel(PAST_FILE)

    current.columns = current.columns.str.lower().str.strip()
    past.columns = past.columns.str.lower().str.strip()

    for col in ['participants', 'budget', 'duration']:
        current[col] = pd.to_numeric(current[col], errors='coerce')
        past[col] = pd.to_numeric(past[col], errors='coerce')

    return current, past


def find_similar_projects(current_value, past_df, column, theme=None,
                          threshold=SIMILARITY_THRESHOLD,
                          min_matches=MIN_MATCHES, max_matches=MAX_MATCHES):
    if FILTER_BY_THEME and theme is not None:
        past_df_filtered = past_df[past_df['theme'].str.strip().str.lower() == theme.strip().lower()].copy()
        if len(past_df_filtered) == 0:
            print(f"  Warning: No past projects with theme '{theme}', using all projects")
            past_df_filtered = past_df.copy()
    else:
        past_df_filtered = past_df.copy()

    lower = current_value * (1 - threshold)
    upper = current_value * (1 + threshold)

    mask = (past_df_filtered[column] >= lower) & (past_df_filtered[column] <= upper)
    similar = past_df_filtered[mask].copy()

    similar['_distance'] = abs(similar[column] - current_value)
    similar = similar.sort_values('_distance')

    if len(similar) < min_matches:
        past_df_temp = past_df_filtered.copy()
        past_df_temp['_distance'] = abs(past_df_temp[column] - current_value)
        past_df_temp = past_df_temp.sort_values('_distance')
        similar = past_df_temp.head(min_matches)

    if max_matches is not None and len(similar) > max_matches:
        similar = similar.head(max_matches)

    similar = similar.drop(columns=['_distance'])

    return similar


def calculate_spread(current_row, similar_projects, metrics):
    spreads = {}

    for metric in metrics:
        current_val = current_row[metric]
        if current_val == 0:
            spreads[metric] = {
                'ratios': [],
                'min_ratio': np.nan,
                'max_ratio': np.nan,
                'mean_ratio': np.nan,
                'std_ratio': np.nan,
                'median_ratio': np.nan,
                'p10': np.nan,
                'p90': np.nan
            }
            continue

        ratios = similar_projects[metric].values / current_val

        spreads[metric] = {
            'ratios': ratios,
            'min_ratio': np.min(ratios),
            'max_ratio': np.max(ratios),
            'mean_ratio': np.mean(ratios),
            'std_ratio': np.std(ratios),
            'median_ratio': np.median(ratios),
            'p10': np.percentile(ratios, 10),
            'p90': np.percentile(ratios, 90)
        }

    return spreads


def analyze_project(current_row, past_df):
    results = {}
    theme = current_row.get('theme', None)

    similar_by_participation = find_similar_projects(
        current_row['participants'], past_df, 'participants', theme=theme
    )
    results['participation_analysis'] = {
        'similar_count': len(similar_by_participation),
        'current_participation': current_row['participants'],
        'spreads': calculate_spread(current_row, similar_by_participation, ['budget', 'duration'])
    }

    similar_by_budget = find_similar_projects(
        current_row['budget'], past_df, 'budget', theme=theme
    )
    results['budget_analysis'] = {
        'similar_count': len(similar_by_budget),
        'current_budget': current_row['budget'],
        'spreads': calculate_spread(current_row, similar_by_budget, ['participants', 'duration'])
    }

    similar_by_duration = find_similar_projects(
        current_row['duration'], past_df, 'duration', theme=theme
    )
    results['duration_analysis'] = {
        'similar_count': len(similar_by_duration),
        'current_duration': current_row['duration'],
        'spreads': calculate_spread(current_row, similar_by_duration, ['participants', 'budget'])
    }

    return results


def create_detailed_results(current_df, past_df):
    """Generate detailed analysis for all current projects."""
    all_results = []

    for idx, row in current_df.iterrows():
        analysis = analyze_project(row, past_df)

        result = {
            'project_index': idx,
            'country': row.get('country', 'N/A'),
            'theme': row.get('theme', 'N/A'),
            'current_participants': row['participants'],
            'current_budget': row['budget'],
            'current_duration': row['duration'],

            'part_similar_count': analysis['participation_analysis']['similar_count'],
            'part_budget_min_ratio': analysis['participation_analysis']['spreads']['budget']['min_ratio'],
            'part_budget_max_ratio': analysis['participation_analysis']['spreads']['budget']['max_ratio'],
            'part_budget_mean_ratio': analysis['participation_analysis']['spreads']['budget']['mean_ratio'],
            'part_budget_p10': analysis['participation_analysis']['spreads']['budget']['p10'],
            'part_budget_p90': analysis['participation_analysis']['spreads']['budget']['p90'],
            'part_duration_min_ratio': analysis['participation_analysis']['spreads']['duration']['min_ratio'],
            'part_duration_max_ratio': analysis['participation_analysis']['spreads']['duration']['max_ratio'],
            'part_duration_mean_ratio': analysis['participation_analysis']['spreads']['duration']['mean_ratio'],
            'part_duration_p10': analysis['participation_analysis']['spreads']['duration']['p10'],
            'part_duration_p90': analysis['participation_analysis']['spreads']['duration']['p90'],

            'budget_similar_count': analysis['budget_analysis']['similar_count'],
            'budget_part_min_ratio': analysis['budget_analysis']['spreads']['participants']['min_ratio'],
            'budget_part_max_ratio': analysis['budget_analysis']['spreads']['participants']['max_ratio'],
            'budget_part_mean_ratio': analysis['budget_analysis']['spreads']['participants']['mean_ratio'],
            'budget_part_p10': analysis['budget_analysis']['spreads']['participants']['p10'],
            'budget_part_p90': analysis['budget_analysis']['spreads']['participants']['p90'],
            'budget_duration_min_ratio': analysis['budget_analysis']['spreads']['duration']['min_ratio'],
            'budget_duration_max_ratio': analysis['budget_analysis']['spreads']['duration']['max_ratio'],
            'budget_duration_mean_ratio': analysis['budget_analysis']['spreads']['duration']['mean_ratio'],
            'budget_duration_p10': analysis['budget_analysis']['spreads']['duration']['p10'],
            'budget_duration_p90': analysis['budget_analysis']['spreads']['duration']['p90'],

            'duration_similar_count': analysis['duration_analysis']['similar_count'],
            'duration_part_min_ratio': analysis['duration_analysis']['spreads']['participants']['min_ratio'],
            'duration_part_max_ratio': analysis['duration_analysis']['spreads']['participants']['max_ratio'],
            'duration_part_mean_ratio': analysis['duration_analysis']['spreads']['participants']['mean_ratio'],
            'duration_part_p10': analysis['duration_analysis']['spreads']['participants']['p10'],
            'duration_part_p90': analysis['duration_analysis']['spreads']['participants']['p90'],
            'duration_budget_min_ratio': analysis['duration_analysis']['spreads']['budget']['min_ratio'],
            'duration_budget_max_ratio': analysis['duration_analysis']['spreads']['budget']['max_ratio'],
            'duration_budget_mean_ratio': analysis['duration_analysis']['spreads']['budget']['mean_ratio'],
            'duration_budget_p10': analysis['duration_analysis']['spreads']['budget']['p10'],
            'duration_budget_p90': analysis['duration_analysis']['spreads']['budget']['p90'],
        }
        all_results.append(result)

    return pd.DataFrame(all_results)


def create_summary_report(detailed_df):
    """Create a summary report of the analysis."""
    summary = {
        'Metric': [],
        'Description': [],
        'Mean': [],
        'Std Dev': [],
        'Min': [],
        'Max': [],
        'Interpretation': []
    }

    metrics_info = [
        ('part_budget_mean_ratio', 'Similar Participation → Budget Ratio',
         'When participation is similar, how budget compares'),
        ('part_duration_mean_ratio', 'Similar Participation → Duration Ratio',
         'When participation is similar, how duration compares'),
        ('budget_part_mean_ratio', 'Similar Budget → Participation Ratio',
         'When budget is similar, how participation compares'),
        ('budget_duration_mean_ratio', 'Similar Budget → Duration Ratio',
         'When budget is similar, how duration compares'),
        ('duration_part_mean_ratio', 'Similar Duration → Participation Ratio',
         'When duration is similar, how participation compares'),
        ('duration_budget_mean_ratio', 'Similar Duration → Budget Ratio',
         'When duration is similar, how budget compares'),
    ]

    for col, desc, interp in metrics_info:
        summary['Metric'].append(col)
        summary['Description'].append(desc)
        summary['Mean'].append(detailed_df[col].mean())
        summary['Std Dev'].append(detailed_df[col].std())
        summary['Min'].append(detailed_df[col].min())
        summary['Max'].append(detailed_df[col].max())
        summary['Interpretation'].append(interp)

    return pd.DataFrame(summary)


def create_predictions_sheet(detailed_df):
    """
    Create a sheet with predicted values for each current project.
    For each metric, multiply current value by mean/min/max ratios.
    Sorted by theme.
    """
    predictions = []

    for _, row in detailed_df.iterrows():
        pred = {
            'project_index': row['project_index'],
            'country': row['country'],
            'theme': row['theme'],

            'current_participants': row['current_participants'],
            'current_budget': row['current_budget'],
            'current_duration': row['current_duration'],

            'participants_pred_min': row['current_participants'] * row['budget_part_min_ratio'],
            'participants_pred_mean': row['current_participants'] * row['budget_part_mean_ratio'],
            'participants_pred_max': row['current_participants'] * row['budget_part_max_ratio'],

            'budget_pred_min': row['current_budget'] * row['part_budget_min_ratio'],
            'budget_pred_mean': row['current_budget'] * row['part_budget_mean_ratio'],
            'budget_pred_max': row['current_budget'] * row['part_budget_max_ratio'],

            'duration_pred_min': row['current_duration'] * row['part_duration_min_ratio'],
            'duration_pred_mean': row['current_duration'] * row['part_duration_mean_ratio'],
            'duration_pred_max': row['current_duration'] * row['part_duration_max_ratio'],
        }
        predictions.append(pred)

    pred_df = pd.DataFrame(predictions)

    pred_df = pred_df.sort_values('theme').reset_index(drop=True)

    return pred_df


def create_hypothesis_comparison(detailed_df):
    comparison = []

    budget_ratios = detailed_df[['part_budget_mean_ratio', 'duration_budget_mean_ratio']].mean().mean()
    comparison.append({
        'Hypothesis': 'Budget: Fully utilized or slightly underused',
        'Expected': '~1.0 or slightly < 1.0',
        'Observed Mean Ratio': f'{budget_ratios:.3f}',
        'Alignment': 'Yes' if 0.8 <= budget_ratios <= 1.1 else 'No'
    })

    part_ratios = detailed_df[['budget_part_mean_ratio', 'duration_part_mean_ratio']].mean().mean()
    comparison.append({
        'Hypothesis': 'Participation: [-10%, +30%]',
        'Expected': '0.9 to 1.3',
        'Observed Mean Ratio': f'{part_ratios:.3f}',
        'Alignment': 'Yes' if 0.7 <= part_ratios <= 1.5 else 'Partial'
    })

    duration_ratios = detailed_df[['part_duration_mean_ratio', 'budget_duration_mean_ratio']].mean().mean()
    comparison.append({
        'Hypothesis': 'Duration: Same or slightly longer',
        'Expected': '~1.0 or slightly > 1.0',
        'Observed Mean Ratio': f'{duration_ratios:.3f}',
        'Alignment': 'Yes' if 0.8 <= duration_ratios <= 1.3 else 'No'
    })

    return pd.DataFrame(comparison)


def style_excel(ws, df, start_row=1, header_fill='4472C4'):
    header_font = Font(bold=True, color='FFFFFF')
    header_fill_style = PatternFill(start_color=header_fill, end_color=header_fill, fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_idx, cell in enumerate(ws[start_row], 1):
        cell.font = header_font
        cell.fill = header_fill_style
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    for row in ws.iter_rows(min_row=start_row + 1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if isinstance(cell.value, float):
                cell.number_format = '0.000'


def export_to_excel(predictions_df, filename):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Predictions"
    for r_idx, row in enumerate(dataframe_to_rows(predictions_df, index=False, header=True), 1):
        for c_idx, val in enumerate(row, 1):
            ws1.cell(row=r_idx, column=c_idx, value=val)
    style_excel(ws1, predictions_df, header_fill='2E7D32')

    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    wb.save(filename)
    print(f"Results exported to {filename}")


def print_summary_report(summary_df, hypothesis_df):
    print("\n" + "=" * 80)
    print("SPREAD ANALYSIS SUMMARY REPORT")
    print("=" * 80)

    print("\n--- METRIC RATIOS (Past / Current) ---\n")
    print(summary_df.to_string(index=False))

    print("\n--- HYPOTHESIS COMPARISON ---\n")
    print(hypothesis_df.to_string(index=False))

    print("\n" + "=" * 80)
    print("INTERPRETATION GUIDE:")
    print("-" * 80)
    print("Ratio = 1.0  : Past and current values are equal")
    print("Ratio > 1.0  : Past projects had higher values than current")
    print("Ratio < 1.0  : Past projects had lower values than current")
    print("P10/P90      : 10th and 90th percentile (likely range)")
    print("=" * 80 + "\n")


def main():
    if RANDOM_SEED is not None:
        np.random.seed(RANDOM_SEED)

    print("Loading data...")
    current_df, past_df = load_data()
    print(f"Current projects: {len(current_df)}")
    print(f"Past projects: {len(past_df)}")

    print("\nAnalyzing projects...")
    detailed_df = create_detailed_results(current_df, past_df)

    print("Generating predictions...")
    predictions_df = create_predictions_sheet(detailed_df)

    print(f"Prediction columns: {list(predictions_df.columns)}")

    print(f"Exporting to {OUTPUT_FILE}...")
    export_to_excel(predictions_df, OUTPUT_FILE)

    print("\nAnalysis complete!")
    return predictions_df


if __name__ == "__main__":
    predictions_df = main()